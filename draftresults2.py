import requests
import json
import openpyxl
from pprint import pprint
import os
import sys


def getFantasyTeams(espn_s2, swid, url):
    fantasyTeamsKey = {}
    r = requests.get(url, cookies={"swid": swid, "espn_s2": espn_s2})
    data = json.loads(r.content)[0]
    for team in data['teams']:
        fantasyTeamsKey[team['id']] = team['location'] + ' ' + team['nickname']
    return fantasyTeamsKey

def getSeasonResults(espn_s2, swid, url, positionsKey, nflTeamsKey):
    playerData = {}
    r = requests.get(url, cookies={"swid": swid, "espn_s2": espn_s2}, params={"view": 'kona_player_info'})
    data = json.loads(r.content)
    for player in data['players']:
        if 'ratings' in player.keys():
            playerData[player['id']] = {}
            playerData[player['id']]['name'] = player['player']['fullName']
            playerData[player['id']]['nflTeam'] = nflTeamsKey[player['player']['proTeamId']]
            playerData[player['id']]['position'] = positionsKey[player['player']['defaultPositionId']]
            playerData[player['id']]['rankOverall'] = int(player['ratings']['0']['totalRanking'])
            playerData[player['id']]['rankPosition'] = int(player['ratings']['0']['positionalRanking'])
            for score in player['player']['stats']:
                if round(float(score['appliedTotal']), 3) == round(float(player['ratings']['0']['totalRating']), 3):
                    playerData[player['id']]['avgScore'] = round(float(score['appliedAverage']), 3)
                    playerData[player['id']]['totalScore'] = round(float(player['ratings']['0']['totalRating']), 3)
            if playerData[player['id']]['avgScore'] == 0:
                playerData[player['id']]['weeksMissed'] = 16
            else:
                playerData[player['id']]['weeksMissed'] = round(16 - playerData[player['id']]['totalScore'] / playerData[player['id']]['avgScore'])
    return playerData

def getDraftResults(espn_s2, swid, url, playerData, fantasyTeamsKey):
    draftData = {}
    draftPositionOrder = {'QB': 1, 'RB': 1, 'WR': 1, 'TE': 1, 'K': 1, 'D/ST': 1, 'HC': 1}
    r = requests.get(url, cookies={"swid": swid, "espn_s2": espn_s2}, params={"view": 'mDraftDetail'})
    data = json.loads(r.content)[0]
    for pick in data['draftDetail']['picks']:
        draftData[pick['playerId']] = playerData[pick['playerId']]
        draftData[pick['playerId']]['pickOverall'] = pick['overallPickNumber']
        draftData[pick['playerId']]['fantasyTeam'] = fantasyTeamsKey[pick['teamId']]
        draftData[pick['playerId']]['pickPosition'] = draftPositionOrder[draftData[pick['playerId']]['position']]
        draftPositionOrder[draftData[pick['playerId']]['position']] += 1
    return draftData

def createSheet(draftData, headersKey):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('Draft+Results')
    for sheet in wb.sheetnames:
        if sheet != 'Draft+Results':
            wb.remove_sheet(wb.get_sheet_by_name(sheet))
    for header in headersKey:
        ws.cell(row=1, column=headersKey[header]).value = header
    for pick in draftData:
        ws.cell(row=draftData[pick]['pickOverall']+1,column=headersKey['Overall Draft Pick']).value = draftData[pick]['pickOverall']
        ws.cell(row=draftData[pick]['pickOverall']+1,column=headersKey['Player Name']).value = draftData[pick]['name']
        ws.cell(row=draftData[pick]['pickOverall']+1,column=headersKey['Fantasy Team']).value = draftData[pick]['fantasyTeam']
        ws.cell(row=draftData[pick]['pickOverall']+1,column=headersKey['Position-Based Draft Pick']).value = draftData[pick]['position'] + '-' + str(draftData[pick]['pickPosition'])
        ws.cell(row=draftData[pick]['pickOverall']+1,column=headersKey['Position-Based Season Finish']).value = draftData[pick]['position'] + '-' + str(draftData[pick]['rankPosition'])
        ws.cell(row=draftData[pick]['pickOverall']+1,column=headersKey['Position']).value = draftData[pick]['position']
        ws.cell(row=draftData[pick]['pickOverall']+1,column=headersKey['Overall Finish']).value = draftData[pick]['rankOverall']
        ws.cell(row=draftData[pick]['pickOverall']+1,column=headersKey['Total Points']).value = draftData[pick]['totalScore']
        ws.cell(row=draftData[pick]['pickOverall']+1,column=headersKey['Number of Weeks Missed']).value = draftData[pick]['weeksMissed']
        ws.cell(row=draftData[pick]['pickOverall']+1,column=headersKey['Average Weekly Scoring']).value = draftData[pick]['avgScore']
    wb.save('Drafts2/' + str(leagueId) + '-' + str(seasonId) + '.xlsx')


if __name__ == "__main__":
    if len(sys.argv) != 5:
        raise ValueError("Error: Please include 4 command line arguments for 2 cookies (espn_s2 and SWID), League ID, and Year")
    espn_s2 = sys.argv[1]
    swid = sys.argv[2]
    leagueId = int(sys.argv[3])
    seasonId = int(sys.argv[4])

    url = 'https://fantasy.espn.com/apis/v3/games/ffl/leagueHistory/'  + str(leagueId) + '?seasonId=' + str(seasonId)
    url2 = 'https://fantasy.espn.com/apis/v3/games/ffl/seasons/' + str(seasonId) + '/segments/0/leagues/'  + str(leagueId) + '?'
    positionsKey = {16: 'D/ST', 14: 'HC', 5: 'K', 1: 'QB', 2: 'RB', 3: 'WR', 4: 'TE', 7: 'K'}
    nflTeamsKey = {0: 'FA', 34: 'Texans', 33: 'Ravens', 30: 'Jaguars', 29: 'Panthers',  28: 'Redskins', 27: 'Buccaneers', 26: 'Seahawks', 25: '49ers', 24: 'Chargers', 23: 'Steelers', 22: 'Cardinals', 21: 'Eagles', 20: 'Jets', 19: 'Giants', 18: 'Saints', 17: 'Patriots', 16: 'Vikings', 15: 'Dolphins', 14: 'Rams', 13: 'Raiders', 12: 'Chiefs', 11: 'Colts', 10: 'Titans', 9: 'Packers', 8: 'Lions', 7: 'Broncos', 6: 'Cowboys', 5: 'Browns', 4: 'Bengals', 3: 'Bears', 2: 'Bills', 1: 'Falcons'}
    headersKey = {'Overall Draft Pick': 1, 'Player Name': 2, 'Position': 3, 'Fantasy Team': 4, 'Pick Rating (1 worst, 10 best)': 5, 'Position-Based Draft Pick': 6, 'Position-Based Season Finish': 7, 'Overall Finish': 8, 'Total Points': 9, 'Number of Weeks Missed': 10, 'Average Weekly Scoring': 11, 'Points in Final 8 Weeks': 12, }

    fantasyTeamsKey = getFantasyTeams(espn_s2, swid, url)
    playerData = getSeasonResults(espn_s2, swid, url2, positionsKey, nflTeamsKey)
    draftData = getDraftResults(espn_s2, swid, url, playerData, fantasyTeamsKey)

    if 'Drafts2/' + str(leagueId) + '-' + str(seasonId) + '.xlsx' in os.listdir():
        raise ValueError("Sheet already exists for league " + str(leagueId) + " in year " + str(seasonId) + ". To create a new sheet, delete the existing one.")
    createSheet(draftData, headersKey)
