import sklearn
from sklearn.linear_model import LogisticRegression, LinearRegression
from sklearn.preprocessing import PolynomialFeatures
from sklearn.svm import SVR
import numpy as np
import openpyxl
import os


def getSamples(positionsDict, variables, headersKey):
    for excel in os.listdir(os.curdir + '/Training/Advanced'):
        wb = openpyxl.load_workbook(filename = os.curdir + '/Training/Advanced/' + excel)
        ws = wb.get_sheet_by_name("Draft+Results")

        row = 2
        pick = ws.cell(row=row, column=1).value
        while pick:
            position = ws.cell(row=row, column=headersKey['Position']).value
            nextRow = [float(ws.cell(row=row, column=headersKey['Position-Based Draft Pick']).value.split('-')[1]), float(ws.cell(row=row, column=headersKey['Position-Based Season Finish']).value.split('-')[1])]
            if nextRow[1] == 0:
                if ws.cell(row=row, column=headersKey['Position']) == 'WR' or 'RB':
                    nextRow[1] = 100
                else:
                    nextRow[1] = 50
            for variable in variables:
                nextRow.append(float(ws.cell(row=row, column=headersKey[variable]).value))
            positionsDict[position]['inputs'].append(nextRow)
            positionsDict[position]['outputs'].append(float(ws.cell(row=row, column=headersKey['Pick Rating (1 worst, 10 best)']).value))
            row += 1
            pick = ws.cell(row=row, column=1).value

def getHeaders(path):
    wb = openpyxl.load_workbook(filename = path)
    ws = wb.get_sheet_by_name("Draft+Results")
    headers = {}
    row = 1
    column = 1
    header = ws.cell(row=row,column=column).value

    while header:
        headers[header] = column
        column += 1
        header = ws.cell(row=row,column=column).value
    
    return headers

def trainSamples(positionsDict):
    for position in positionsDict:
        svRegressor = SVR()
        svRegressor.fit(np.asarray(positionsDict[position]['inputs']), np.asarray(positionsDict[position]['outputs']))
        positionsDict[position]['regressor'] = svRegressor

def evaluate(positionsDict, variables, headersKey):
    for excel in os.listdir(os.curdir + '/Drafts/Advanced'):
        wb = openpyxl.load_workbook(filename = os.curdir + '/Drafts/Advanced/' + excel)
        ws = wb.get_sheet_by_name("Draft+Results")

        row = 2
        pick = ws.cell(row=row, column=1).value
        while pick:
            position = ws.cell(row=row, column=headersKey['Position']).value
            sample = [float(ws.cell(row=row, column=headersKey['Position-Based Draft Pick']).value.split('-')[1]), float(ws.cell(row=row, column=headersKey['Position-Based Season Finish']).value.split('-')[1])]
            if sample[1] == 0:
                if ws.cell(row=row, column=headersKey['Position']) == 'WR' or 'RB':
                    sample[1] = 100
                else:
                    sample[1] = 50
            for variable in variables:
                sample.append(float(ws.cell(row=row, column=headersKey[variable]).value))
            ws.cell(row=row, column = headersKey['Pick Rating (1 worst, 10 best)']).value = str(list(positionsDict[position]['regressor'].predict(np.asarray([sample])))[0])
            row += 1
            pick = ws.cell(row=row, column=1).value

        wb.save(os.curdir + '/Fitted/Advanced10/' + excel)


if __name__ == "__main__":
    positionsDict = {'D/ST': {'inputs': [], 'outputs': []}, 'HC': {'inputs': [], 'outputs': []}, 'K': {'inputs': [], 'outputs': []}, 'QB': {'inputs': [], 'outputs': []}, 'RB': {'inputs': [], 'outputs': []}, 'WR': {'inputs': [], 'outputs': []}, 'TE': {'inputs': [], 'outputs': []}}
    variables = ['Overall Draft Pick', 'Overall Finish', 'Total Points', 'Number of Weeks Missed', 'Average Weekly Scoring']
    
    for excel in os.listdir(os.curdir + '/Fitted/Advanced10'):
        os.remove(os.curdir + '/Fitted/Advanced10/' + excel)
    if len(os.listdir(os.curdir + '/Training/Advanced')) == 0:
        raise ValueError("No Training Data in Training/Advanced")

    headersKey = getHeaders(os.curdir + '/Training/Advanced/' + os.listdir(os.curdir + '/Training/Advanced')[0])
    getSamples(positionsDict, variables, headersKey)

    trainSamples(positionsDict)

    evaluate(positionsDict, variables, headersKey)
