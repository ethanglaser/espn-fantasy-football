import sklearn
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import MinMaxScaler
from sklearn.model_selection import train_test_split
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error, median_absolute_error
import numpy as np
import openpyxl
import os
from colour import Color
import math
import pandas as pd
from matplotlib import pyplot as plt
import xgboost

def get_data():
    data = []
    for excel in os.listdir(os.curdir + '/Training'):
        current = pd.read_excel(os.curdir + '/Training/' + excel)
        current = initial_processing(current)
        data.append(current)
    data = pd.concat(data)
    data = data.dropna()
    return data

def initial_processing(current):
    if "Points in Final 8 Weeks" in current.columns:
        current = current.drop(columns=["Points in Final 8 Weeks"])
    current = current.dropna()
    current = pd.get_dummies(current, columns=["Position"], prefix="Position_")
    current["Position-Based Draft Pick"] = current["Position-Based Draft Pick"].str.extract(r"[A-Z]+-(\d+)")
    current["Position-Based Season Finish"] = current["Position-Based Season Finish"].str.extract(r"[A-Z]+-(\d+)")    
    current = current.replace({"Position-Based Season Finish": 0}, 100)
    return current

def fit_model(data, model, seed):
    train, test = train_test_split(data, test_size=0.2, random_state=seed)
    train_x = train.drop(columns=["Pick Rating (1 worst, 10 best)"])
    train_y = train["Pick Rating (1 worst, 10 best)"]
    print(train_x)
    print(train_y)
    model.fit(train_x, train_y)

    return model, train_x, train_y, test.drop(columns=["Pick Rating (1 worst, 10 best)"]), test["Pick Rating (1 worst, 10 best)"]

def evaluate_model(model, test_x, test_y):
    pred_y = model.predict(test_x)
    r2 = r2_score(test_y, pred_y)
    rmse = np.sqrt(mean_squared_error(test_y, pred_y))
    mae = mean_absolute_error(test_y, pred_y)
    medae = median_absolute_error(test_y, pred_y)
    return r2, rmse, mae, medae

def evaluate(positionsDict, variables, headersKey):
    for excel in os.listdir(os.curdir + '/Drafts'):
        teamNames = set()
        leaderData = []
        wb = openpyxl.load_workbook(filename = os.curdir + '/Drafts/' + excel)
        ws = wb.get_sheet_by_name("Draft+Results")
        row = 2
        pick = ws.cell(row=row, column=1).value
        while pick:
            teamNames.add(ws.cell(row=row, column=headersKey['Fantasy Team']).value)
            position = ws.cell(row=row, column=headersKey['Position']).value
            sample = [float(ws.cell(row=row, column=headersKey['Position-Based Draft Pick']).value.split('-')[1]), float(ws.cell(row=row, column=headersKey['Position-Based Season Finish']).value.split('-')[1])]
            if sample[1] == 0:
                if ws.cell(row=row, column=headersKey['Position']) == 'WR' or 'RB':
                    sample[1] = 100
                else:
                    sample[1] = 50
            for variable in variables:
                if ws.cell(row=row, column=headersKey[variable]).value != None:
                    sample.append(float(ws.cell(row=row, column=headersKey[variable]).value))
            ws.cell(row=row, column = headersKey['Pick Rating (1 worst, 10 best)']).value = str(round(list(positionsDict[position]['regressor'].predict(np.asarray([sample])))[0], 4))
            newLeader = {}
            newLeader['name'] = ws.cell(row=row, column=headersKey['Player Name']).value
            newLeader['rating'] = float(ws.cell(row=row, column=headersKey['Pick Rating (1 worst, 10 best)']).value)
            newLeader['team'] = ws.cell(row=row, column=headersKey['Fantasy Team']).value
            leaderData.append(newLeader)
            row += 1
            pick = ws.cell(row=row, column=1).value
        teamSheets(teamNames, headersKey, wb)
        leaderboards(teamNames, leaderData, wb)
        wb.save(os.curdir + '/Fitted/' + excel)

def teamSheets(teams, headers, wb):
    teamPicks = {}
    namesKey = {}
    colors = list(Color("red").range_to(Color("yellow"),6))[:-1] + list(Color("yellow").range_to(Color("green"),5))
    for index, team in enumerate(teams):
        teamPicks[team] = 2
        try:
            wb.create_sheet(team)
            namesKey[team] = team
        except:
            name = 'Team' + str(index + 1)
            wb.create_sheet(name)
            namesKey[team] = name
        ws2 = wb.get_sheet_by_name(namesKey[team])
        for header in headers:
            ws2.cell(row=1, column=headers[header]).value = header
    ws = wb.get_sheet_by_name("Draft+Results")
    row = 2
    pick = ws.cell(row=row, column=1).value
    while pick:
        team = ws.cell(row=row, column=headersKey['Fantasy Team']).value
        ws2 = wb.get_sheet_by_name(namesKey[team])
        for header in headers:
            ws2.cell(row=teamPicks[team], column=headers[header]).value = ws.cell(row=row, column=headers[header]).value
        colorIndex = math.floor(float(ws2.cell(row=teamPicks[team], column=headers['Pick Rating (1 worst, 10 best)']).value))
        if colorIndex > 9:
            colorIndex = 9
        if colorIndex < 0:
            colorIndex = 0
        ws2.cell(row=teamPicks[team], column=headers['Pick Rating (1 worst, 10 best)']).fill = openpyxl.styles.PatternFill(start_color=str(colors[colorIndex].hex_l)[1:], end_color=str(colors[colorIndex].hex_l)[1:], fill_type = 'solid')
        teamPicks[team] += 1
        row += 1
        pick = ws.cell(row=row, column=1).value
    for team in teams:
        ws2 = wb.get_sheet_by_name(namesKey[team])
        ws2.delete_cols(headers['Fantasy Team'])
        ws2.column_dimensions['B'].width = 25.0
        ws2.column_dimensions['D'].width = 25.0

def leaderboards(teams, data, wb):
    wb.create_sheet("Leaderboards")
    ws = wb.get_sheet_by_name("Leaderboards")
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20

    teamAverages = {}
    for team in teams:
        teamAverages[team] = []
        for player in data:
            if team == player['team']:
                teamAverages[team].append(player['rating'])
        teamAverages[team] = sum(teamAverages[team])/len(teamAverages[team])

    ws.cell(row=1, column=1).value = 'Team Average'
    ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
    row = 2
    for team in sorted(teamAverages.items(), key=lambda x: x[1], reverse=True):
        ws.cell(row=row, column=1).value = team[0]
        ws.cell(row=row, column=2).value = teamAverages[team[0]]
        row += 1
    row += 1
    ws.cell(row=row, column=1).value = "Best Picks"    
    ws.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
    for player in sorted(data, key=lambda i: i['rating'], reverse=True)[:10]:
        row += 1
        ws.cell(row=row, column=3).value = player['rating']
        ws.cell(row=row, column=1).value = player['name']
        ws.cell(row=row, column=2).value = player['team']
    
    row += 2
    ws.cell(row=row, column=1).value = "Worst Picks"    
    ws.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
    for player in sorted(data, key=lambda i: i['rating'])[:10]:
        row += 1
        ws.cell(row=row, column=3).value = player['rating']
        ws.cell(row=row, column=1).value = player['name']
        ws.cell(row=row, column=2).value = player['team']
 

if __name__ == "__main__":    
    for excel in os.listdir(os.curdir + '/Fitted'):
        os.remove(os.curdir + '/Fitted/' + excel)
    if len(os.listdir(os.curdir + '/Training')) == 0:
        raise ValueError("No Training Data in Training Directory")

    training = get_data()

    model = xgboost.XGBRegressor()
    seed = 0

    model, train_x, train_y, test_x, test_y = fit_model(training.select_dtypes(['number']), model, seed)
    print(evaluate_model(model, test_x, test_y))

    #evaluate(positionsDict, variables, headersKey)
