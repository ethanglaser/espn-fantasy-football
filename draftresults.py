import requests
import json
import openpyxl
from pprint import pprint
import os
import sys


def getFantasyTeams(espn_s2, swid, url):
    fantasyTeamsKey = {}
    r = requests.get(url, cookies={"swid": swid, "espn_s2": espn_s2})
    data = json.loads(r.content)
    print(data.keys())
    for team in data['teams']:
        fantasyTeamsKey[team['id']] = team['location'] + ' ' + team['nickname']
    return fantasyTeamsKey

def getSeasonResults(espn_s2, swid, url, positionsKey, nflTeamsKey, leagueId, seasonId):
    playerData = {}
    url = 'https://fantasy.espn.com/apis/v3/games/ffl/seasons/' + str(seasonId) + '/segments/0/leagues/'  + str(leagueId) + '?view=kona_player_info'
    r = requests.get(url, cookies={"swid": swid, "espn_s2": espn_s2}, headers={'x-fantasy-filter': '{"players": {"sortAppliedStatTotal":{"sortAsc":false,"sortPriority":2,"value":"00' + str(seasonId) + '"}}}'})
    #r = requests.get(url, cookies={"swid": swid, "espn_s2": espn_s2}, params={"view": 'kona_player_info'}, headers={"x-fantasy-filter": '{"players":{"filterSlotIds":{"value":[0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,23,24]},"sortAppliedStatTotal":{"sortAsc":false,"sortPriority":2,"value":"002020"},"sortAppliedStatTotalForScoringPeriodId":null,"sortStatId":null,"sortStatIdForScoringPeriodId":null,"sortPercOwned":{"sortPriority":3,"sortAsc":false},"filterRanksForSlotIds":{"value":[0,2,4,6,17,16]},"filterStatsForTopScoringPeriodIds":{"value":2,"additionalValue":["002020","102020","002019","022020"]}}}'})
    data = json.loads(r.content)
    print(data.keys())
    #print(data[].keys())
    print(len(data['players']))
    for player in data['players']:
        #print(player['player']['fullName'])
        #pprint(player['player']['stats'])
        if 'ratings' in player.keys():
            playerData[player['id']] = {}
            playerData[player['id']]['Player Name'] = player['player']['fullName']
            playerData[player['id']]['nflTeam'] = nflTeamsKey[player['player']['proTeamId']]
            playerData[player['id']]['Position'] = positionsKey[player['player']['defaultPositionId']]
            playerData[player['id']]['Overall Finish'] = int(player['ratings']['0']['totalRanking'])
            playerData[player['id']]['rankPosition'] = int(player['ratings']['0']['positionalRanking'])
            '''print(player['player']['fullName'])
            print()
            print(player['ratings'])
            print()
            for score in player['player']['stats']:
                print(score)
                if round(float(score['appliedTotal']), 3) == round(float(player['ratings']['0']['totalRating']), 3) and 'appliedAverage' in score.keys():
                    playerData[player['id']]['Average Weekly Scoring'] = round(float(score['appliedAverage']), 3)
                    playerData[player['id']]['Total Points'] = round(float(player['ratings']['0']['totalRating']), 3)
            if playerData[player['id']]['Average Weekly Scoring'] == 0:
                playerData[player['id']]['Number of Weeks Missed'] = 16
            else:'''
            try:
                if player['player']['fullName'] == "Rondale Moore" or player['player']['fullName'] == "Tom Brady":
                    for a in player['player']['stats']:
                        pprint(a['appliedAverage'], a['appliedTotal'])
                playerData[player['id']]['Average Weekly Scoring'] = round(float(player['player']['stats'][2]['appliedAverage']), 3)
                playerData[player['id']]['Total Points'] = round(float(player['ratings']['0']['totalRating']), 3)
                if playerData[player['id']]['Position'] == 'HC' or playerData[player['id']]['Position'] == 'D/ST':
                    playerData[player['id']]['Number of Weeks Missed'] = 0
                elif playerData[player['id']]['Average Weekly Scoring'] == 0:
                    playerData[player['id']]['Number of Weeks Missed'] = 16
                else:
                    playerData[player['id']]['Number of Weeks Missed'] = round(16 - playerData[player['id']]['Total Points'] / playerData[player['id']]['Average Weekly Scoring'])

            except:
                print(player['player'])
    return playerData

def getDraftResults(espn_s2, swid, url, playerData, fantasyTeamsKey):
    draftData = {}
    draftPositionOrder = {'QB': 1, 'RB': 1, 'WR': 1, 'TE': 1, 'K': 1, 'D/ST': 1, 'HC': 1}
    r = requests.get(url, cookies={"swid": swid, "espn_s2": espn_s2}, params={"view": 'mDraftDetail'})
    data = json.loads(r.content)
    print(len(playerData))
    for pick in data['draftDetail']['picks']:
        draftData[pick['playerId']] = playerData[pick['playerId']]
        draftData[pick['playerId']]['Overall Draft Pick'] = pick['overallPickNumber']
        draftData[pick['playerId']]['Fantasy Team'] = fantasyTeamsKey[pick['teamId']]
        draftData[pick['playerId']]['pickPosition'] = draftPositionOrder[draftData[pick['playerId']]['Position']]
        draftPositionOrder[draftData[pick['playerId']]['Position']] += 1

    return draftData

def createSheet(draftData, headersKey):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('Draft+Results')
    ws.column_dimensions['B'].width = 25.0
    ws.column_dimensions['D'].width = 25.0
    ws.column_dimensions['E'].width = 25.0
    for sheet in wb.sheetnames:
        if sheet != 'Draft+Results':
            wb.remove(wb.get_sheet_by_name(sheet))
    for header in headersKey:
        ws.cell(row=1, column=headersKey[header]).value = header
    for pick in draftData:
        ws.cell(row=draftData[pick]['Overall Draft Pick']+1,column=headersKey['Position-Based Draft Pick']).value = draftData[pick]['Position'] + '-' + str(draftData[pick]['pickPosition'])
        ws.cell(row=draftData[pick]['Overall Draft Pick']+1,column=headersKey['Position-Based Season Finish']).value = draftData[pick]['Position'] + '-' + str(draftData[pick]['rankPosition'])
        for key in headersKey.keys():
            if key in draftData[pick].keys():
                ws.cell(row=draftData[pick]['Overall Draft Pick']+1,column=headersKey[key]).value = draftData[pick][key]
    wb.save('Drafts/' + str(leagueId) + '-' + str(seasonId) + '.xlsx')


if __name__ == "__main__":
    if len(sys.argv) != 5:
        raise ValueError("Error: Please include 4 command line arguments for 2 cookies (espn_s2 and SWID), League ID, and Year")
    espn_s2 = sys.argv[1]
    swid = sys.argv[2]
    leagueId = int(sys.argv[3])
    seasonId = int(sys.argv[4])

    url = 'https://fantasy.espn.com/apis/v3/games/ffl/leagueHistory/'  + str(leagueId) + '?seasonId=' + str(seasonId)
    url2 = 'https://fantasy.espn.com/apis/v3/games/ffl/seasons/' + str(seasonId) + '/segments/0/leagues/'  + str(leagueId) + '?'
    positionsKey = {16: 'D/ST', 14: 'HC', 5: 'K', 1: 'QB', 2: 'RB', 3: 'WR', 4: 'TE', 7: 'K', 9: 'RB'}
    nflTeamsKey = {0: 'FA', 34: 'Texans', 33: 'Ravens', 30: 'Jaguars', 29: 'Panthers',  28: 'Redskins', 27: 'Buccaneers', 26: 'Seahawks', 25: '49ers', 24: 'Chargers', 23: 'Steelers', 22: 'Cardinals', 21: 'Eagles', 20: 'Jets', 19: 'Giants', 18: 'Saints', 17: 'Patriots', 16: 'Vikings', 15: 'Dolphins', 14: 'Rams', 13: 'Raiders', 12: 'Chiefs', 11: 'Colts', 10: 'Titans', 9: 'Packers', 8: 'Lions', 7: 'Broncos', 6: 'Cowboys', 5: 'Browns', 4: 'Bengals', 3: 'Bears', 2: 'Bills', 1: 'Falcons'}
    headersKey = {'Overall Draft Pick': 1, 'Player Name': 2, 'Position': 3, 'Fantasy Team': 4, 'Pick Rating (1 worst, 10 best)': 5, 'Position-Based Draft Pick': 6, 'Position-Based Season Finish': 7, 'Overall Finish': 8, 'Total Points': 9, 'Number of Weeks Missed': 10, 'Average Weekly Scoring': 11}

    fantasyTeamsKey = getFantasyTeams(espn_s2, swid, url2)
    playerData = getSeasonResults(espn_s2, swid, url2, positionsKey, nflTeamsKey, leagueId, seasonId)
    draftData = getDraftResults(espn_s2, swid, url2, playerData, fantasyTeamsKey)

    if 'Drafts' not in os.listdir():
        os.mkdir('Drafts')
    if str(leagueId) + '-' + str(seasonId) + '.xlsx' in os.listdir('Drafts/'):
        raise ValueError("Sheet already exists for league " + str(leagueId) + " in year " + str(seasonId) + ". To create a new sheet, delete the existing one.")
    createSheet(draftData, headersKey)
